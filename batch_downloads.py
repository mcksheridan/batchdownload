import openpyxl

from openpyxl import load_workbook
wb = load_workbook('sourcelist.xlsx')
ws = wb['List']
rows = ws.max_row

import re, requests

for i in range(rows):
    i += 1
    url = ws[f'A{i}'].value #URL column entry
    match = re.compile(r'\.[a-zA-z][a-zA-z][a-zA-z]$') #Confirm the URL has a three-letter file extension.
    if match.search(url):
        print(f'Downloading {url}...')
        target = ws[f'B{i}'].value #The name of the output file, sans extension.
        extension = url[-4:]
        resp = requests.get(url)
        output = open(f'sources/{target}{extension}', 'wb') #Save the file in the "sources" folder with the correct extension.
        output.write(resp.content)
        output.close()
        print(f'Saved as {target}{extension}')
    else:
        continue

print('Downloads complete')